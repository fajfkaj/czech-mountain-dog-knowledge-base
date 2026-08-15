"""Fáze 1 — archivní inventář (spec §2.16, master prompt bod 5).

Projde všechna zdrojová PDF (přes src.paths, tvrdě vyloučen podadresář
s novějšími čísly), zjistí metadata z pdfinfo + názvu souboru, a zapíše:
  - 00_inventar/inventar.json  (Inventar, Casova_osa, Mezery)
  - 03_vystupy/01_ARCHIVNI_INVENTAR.xlsx  (tři listy, freeze_panes + autofilter)

Spouštět jako modul z kořene projektu: `python3 -m src.inventar`
"""
from __future__ import annotations

import json
import re
import subprocess
from pathlib import Path

import yaml
from openpyxl import Workbook

from src.paths import find_source_pdfs

PROJEKT_ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = PROJEKT_ROOT / "config" / "config.yaml"
INVENTAR_JSON = PROJEKT_ROOT / "00_inventar" / "inventar.json"
INVENTAR_XLSX = PROJEKT_ROOT / "03_vystupy" / "01_ARCHIVNI_INVENTAR.xlsx"

NEUVEDENO = "neuvedeno / nelze jednoznačně určit"

# 45-1984-06_zpravodaj_c_1.pdf / 88-1999_sprezeni_c_41.pdf (bez měsíce) /
# 68-1992-11_sprezeni_special.pdf / 63-1992-01_chovatelsky_zapisni_rad.pdf
NAZEV_RE = re.compile(
    r"^(?P<prefix>\d+)-(?P<rok>\d{4})(?:-(?P<mesic>\d{2}))?_(?P<zbytek>.+)\.pdf$",
    re.IGNORECASE,
)
CISLO_RE = re.compile(r"_c_(?P<cislo>\d+)$", re.IGNORECASE)


def nacti_config(cesta: Path = CONFIG_PATH) -> dict:
    with open(cesta, encoding="utf-8") as f:
        return yaml.safe_load(f)


def _pdfinfo_pocet_stran(pdf: Path) -> tuple[int | None, str | None]:
    try:
        vysledek = subprocess.run(
            ["pdfinfo", str(pdf)], capture_output=True, text=True, timeout=30
        )
    except (subprocess.TimeoutExpired, OSError) as e:
        return None, f"chyba: pdfinfo selhalo ({e})"
    if vysledek.returncode != 0:
        return None, f"chyba: pdfinfo returncode {vysledek.returncode}: {vysledek.stderr.strip()[:200]}"
    for radek in vysledek.stdout.splitlines():
        if radek.startswith("Pages:"):
            try:
                return int(radek.split(":", 1)[1].strip()), None
            except ValueError:
                return None, "chyba: nečitelný počet stran z pdfinfo"
    return None, "chyba: pdfinfo neobsahovalo Pages"


def _dostupnost_textu(pdf: Path) -> tuple[str, bool]:
    """Vrátí (Dostupnost_textu, Nutnost_OCR)."""
    try:
        vysledek = subprocess.run(
            ["pdftotext", str(pdf), "-"], capture_output=True, text=True, timeout=60
        )
    except (subprocess.TimeoutExpired, OSError) as e:
        return f"chyba: pdftotext selhalo ({e})", True
    if vysledek.returncode != 0:
        return f"chyba: pdftotext returncode {vysledek.returncode}", True
    znaku = len(vysledek.stdout.strip())
    if znaku < 50:
        return "skenované", True
    return "textové", False


def _rozparsuj_nazev(nazev: str) -> dict:
    m = NAZEV_RE.match(nazev)
    if not m:
        return {
            "prefix": None,
            "rok": None,
            "mesic": None,
            "cislo_zpravodaje": NEUVEDENO,
            "typ_dokumentu": NEUVEDENO,
        }
    prefix = int(m.group("prefix"))
    rok = int(m.group("rok"))
    mesic = int(m.group("mesic")) if m.group("mesic") else None
    zbytek = m.group("zbytek").lower()

    if "special" in zbytek:
        typ = "zvláštní číslo"
        mesic_str = f"{mesic:02d}" if mesic else "??"
        cislo = f"zvláštní ({rok}-{mesic_str})"
    elif "chovatelsky_zapisni_rad" in zbytek:
        typ = "chovatelský a zápisní řád"
        mesic_str = f"{mesic:02d}" if mesic else "??"
        cislo = f"Chovatelský a zápisní řád ({rok}-{mesic_str})"
    else:
        typ = "řadové číslo"
        cm = CISLO_RE.search(zbytek)
        cislo = cm.group("cislo") if cm else NEUVEDENO

    return {
        "prefix": prefix,
        "rok": rok,
        "mesic": mesic,
        "cislo_zpravodaje": cislo,
        "typ_dokumentu": typ,
    }


def sestav_inventar(pdfs: list[Path]) -> list[dict]:
    zaznamy = []
    for pdf in pdfs:
        meta = _rozparsuj_nazev(pdf.name)
        pocet_stran, chyba_stran = _pdfinfo_pocet_stran(pdf)

        if chyba_stran:
            dostupnost = chyba_stran
            nutnost_ocr = False
            problemy = chyba_stran
        else:
            dostupnost, nutnost_ocr = _dostupnost_textu(pdf)
            problemy = "" if not dostupnost.startswith("chyba") else dostupnost

        zaznamy.append(
            {
                "Nazev_souboru": pdf.name,
                "Prefix": meta["prefix"],
                "Cislo_zpravodaje": meta["cislo_zpravodaje"],
                "Rok": meta["rok"],
                "Mesic": meta["mesic"] if meta["mesic"] else NEUVEDENO,
                "Pocet_stran": pocet_stran if pocet_stran is not None else NEUVEDENO,
                "Dostupnost_textu": dostupnost,
                "Nutnost_OCR": "ano" if nutnost_ocr else "ne",
                "Problemy_s_kvalitou": problemy or NEUVEDENO,
                "Typ_dokumentu": meta["typ_dokumentu"],
            }
        )
    zaznamy.sort(key=lambda z: (z["Prefix"] is None, z["Prefix"]))
    return zaznamy


def sestav_casovou_osu(zaznamy: list[dict]) -> dict[str, list[str]]:
    osa: dict[str, list[str]] = {}
    for z in zaznamy:
        rok = z["Rok"]
        if rok is None:
            continue
        osa.setdefault(str(rok), []).append(z["Cislo_zpravodaje"])
    for rok in osa:
        osa[rok].sort()
    return dict(sorted(osa.items()))


def sestav_mezery(casova_osa: dict[str, list[str]]) -> list[dict]:
    roky_s_daty = sorted(int(r) for r in casova_osa)
    if not roky_s_daty:
        return []
    min_rok, max_rok = roky_s_daty[0], roky_s_daty[-1]
    mezery = []
    for rok in range(min_rok, max_rok + 1):
        if str(rok) not in casova_osa:
            mezery.append(
                {
                    "Rok": rok,
                    "Poznamka": "chybí číslo v archivu pro tento rok — mezera v datech, "
                    "ne domněnka o neexistenci čísla",
                }
            )
    return mezery


def zapis_json(zaznamy, casova_osa, mezery) -> None:
    INVENTAR_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(INVENTAR_JSON, "w", encoding="utf-8") as f:
        json.dump(
            {"Inventar": zaznamy, "Casova_osa": casova_osa, "Mezery": mezery},
            f,
            ensure_ascii=False,
            indent=2,
        )


def zapis_xlsx(zaznamy, casova_osa, mezery) -> None:
    INVENTAR_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Inventar"
    sloupce = list(zaznamy[0].keys()) if zaznamy else []
    ws1.append(sloupce)
    for z in zaznamy:
        ws1.append([z[s] for s in sloupce])
    ws1.freeze_panes = "A2"
    if zaznamy:
        ws1.auto_filter.ref = ws1.dimensions

    ws2 = wb.create_sheet("Casova_osa")
    ws2.append(["Rok", "Cisla_vydana_v_roce"])
    for rok, cisla in casova_osa.items():
        ws2.append([int(rok), ", ".join(str(c) for c in cisla)])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    ws3 = wb.create_sheet("Mezery")
    ws3.append(["Rok", "Poznamka"])
    for m in mezery:
        ws3.append([m["Rok"], m["Poznamka"]])
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions

    wb.save(INVENTAR_XLSX)


def main() -> None:
    cfg = nacti_config()
    pdfs = find_source_pdfs(cfg["zdroj_koren"], cfg["zdroj_filtr"], cfg["vylouceno"])
    zaznamy = sestav_inventar(pdfs)
    casova_osa = sestav_casovou_osu(zaznamy)
    mezery = sestav_mezery(casova_osa)
    zapis_json(zaznamy, casova_osa, mezery)
    zapis_xlsx(zaznamy, casova_osa, mezery)

    print(f"Nalezeno souborů: {len(zaznamy)}")
    print(f"Celkem stran: {sum(z['Pocet_stran'] for z in zaznamy if isinstance(z['Pocet_stran'], int))}")
    print(f"Chybějící ročníky: {[m['Rok'] for m in mezery]}")
    chyby = [z for z in zaznamy if str(z["Dostupnost_textu"]).startswith("chyba")]
    if chyby:
        print(f"POZOR — {len(chyby)} souborů s chybou: {[z['Nazev_souboru'] for z in chyby]}")
    print(f"Zapsáno: {INVENTAR_JSON}")
    print(f"Zapsáno: {INVENTAR_XLSX}")


if __name__ == "__main__":
    main()
